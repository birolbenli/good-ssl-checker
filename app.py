import ssl
import socket
import datetime
import threading
import time
import uuid
import json
import os
import smtplib
import schedule
import email.mime.text
import email.mime.multipart
import io
import cryptography.x509
from concurrent.futures import ThreadPoolExecutor, as_completed
from flask import Flask, render_template, request, jsonify, session, send_file, redirect, url_for, flash
import requests
import sqlite3
from contextlib import contextmanager
from openpyxl import load_workbook
from functools import wraps

app = Flask(__name__)
app.secret_key = 'ssl-checker-secret-key-2025'

# Simple authentication credentials
USERNAME = 'admin'
PASSWORD = 'ssl123'

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'logged_in' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# Database helper functions
@contextmanager
def get_db_connection():
    # Ensure instance directory exists
    instance_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'instance')
    os.makedirs(instance_dir, exist_ok=True)
    db_path = os.path.join(instance_dir, 'ssl_tracker.db')
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    try:
        yield conn
    finally:
        conn.close()

def init_default_settings(conn):
    """Initialize default settings in the database"""
    default_settings = [
        # SMTP Settings
        ('smtp_server', 'smtp.gmail.com', 'smtp', 'string', 'SMTP server hostname'),
        ('smtp_port', '587', 'smtp', 'number', 'SMTP server port'),
        ('smtp_username', '', 'smtp', 'string', 'SMTP username'),
        ('smtp_password', '', 'smtp', 'password', 'SMTP password'),
        ('smtp_from_email', '', 'smtp', 'string', 'From email address'),
        ('smtp_use_tls', 'true', 'smtp', 'boolean', 'Use TLS encryption'),
        
        # Default Email Settings
        ('default_to_email', '', 'email', 'string', 'Default recipient email address'),
        
        # Default Slack Settings
        ('default_slack_webhook', '', 'slack', 'string', 'Default Slack webhook URL'),
        ('slack_channel', '#ssl-alerts', 'slack', 'string', 'Default Slack channel'),
        ('slack_username', 'SSL Checker Bot', 'slack', 'string', 'Slack bot username'),
        ('slack_icon_emoji', ':warning:', 'slack', 'string', 'Slack bot emoji'),
        
        # Notification Settings
        ('notifications_enabled', 'true', 'notifications', 'boolean', 'Enable notifications'),
        ('schedule_time', '09:00', 'notifications', 'string', 'Daily notification time (HH:MM)'),
        ('expiry_thresholds', '7,15,30', 'notifications', 'string', 'Days before expiry to send alerts (comma separated)'),
        ('email_enabled', 'true', 'notifications', 'boolean', 'Enable email notifications'),
        ('slack_enabled', 'true', 'notifications', 'boolean', 'Enable Slack notifications'),
    ]
    
    for key, value, category, type_name, description in default_settings:
        # Check if setting already exists
        existing = conn.execute('SELECT id FROM settings WHERE key = ?', (key,)).fetchone()
        if not existing:
            conn.execute('''
                INSERT INTO settings (key, value, category, type, description)
                VALUES (?, ?, ?, ?, ?)
            ''', (key, value, category, type_name, description))

def get_setting(key, default_value=None):
    """Get a setting value from database"""
    try:
        with get_db_connection() as conn:
            result = conn.execute('SELECT value FROM settings WHERE key = ?', (key,)).fetchone()
            return result['value'] if result else default_value
    except Exception as e:
        print(f"Error getting setting {key}: {e}")
        return default_value

def set_setting(key, value):
    """Set a setting value in database"""
    try:
        with get_db_connection() as conn:
            conn.execute('''
                INSERT OR REPLACE INTO settings (key, value, updated_at)
                VALUES (?, ?, ?)
            ''', (key, value, datetime.datetime.now().isoformat()))
            conn.commit()
            return True
    except Exception as e:
        print(f"Error setting {key}: {e}")
        return False

def get_all_settings():
    """Get all settings grouped by category"""
    try:
        with get_db_connection() as conn:
            results = conn.execute('''
                SELECT key, value, category, type, description 
                FROM settings 
                ORDER BY category, key
            ''').fetchall()
            
            settings = {}
            for row in results:
                category = row['category']
                if category not in settings:
                    settings[category] = {}
                
                value = row['value']
                # Convert boolean strings
                if row['type'] == 'boolean':
                    value = value.lower() == 'true'
                elif row['type'] == 'number':
                    try:
                        value = int(value)
                    except:
                        value = 0
                
                settings[category][row['key']] = {
                    'value': value,
                    'type': row['type'],
                    'description': row['description']
                }
            
            return settings
    except Exception as e:
        print(f"Error getting all settings: {e}")
        return {}

def init_db():
    """Initialize the database with required tables"""
    with get_db_connection() as conn:
        conn.execute('''
            CREATE TABLE IF NOT EXISTS domain (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        conn.execute('''
            CREATE TABLE IF NOT EXISTS subdomain (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                domain_id INTEGER,
                dns TEXT,
                source_dns TEXT,
                proxied TEXT,
                record_type TEXT DEFAULT 'A',
                ip TEXT,
                nat TEXT,
                port INTEGER DEFAULT 443,
                owner TEXT,
                certificate TEXT,
                expire_date DATE,
                days_left INTEGER,
                certificate_renewal_status TEXT,
                last_checked DATETIME,
                email_notification INTEGER DEFAULT 1,
                slack_notification INTEGER DEFAULT 1,
                custom_email TEXT,
                custom_slack_webhook TEXT,
                FOREIGN KEY (domain_id) REFERENCES domain (id)
            )
        ''')
        
        conn.execute('''
            CREATE TABLE IF NOT EXISTS settings (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                key TEXT UNIQUE NOT NULL,
                value TEXT,
                category TEXT,
                type TEXT DEFAULT 'string',
                description TEXT,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Add notification columns if they don't exist
        try:
            conn.execute('ALTER TABLE subdomain ADD COLUMN email_notification INTEGER DEFAULT 1')
        except sqlite3.OperationalError:
            pass  # Column already exists
            
        try:
            conn.execute('ALTER TABLE subdomain ADD COLUMN slack_notification INTEGER DEFAULT 1')
        except sqlite3.OperationalError:
            pass  # Column already exists
            
        try:
            conn.execute('ALTER TABLE subdomain ADD COLUMN custom_email TEXT')
        except sqlite3.OperationalError:
            pass  # Column already exists
            
        try:
            conn.execute('ALTER TABLE subdomain ADD COLUMN custom_slack_webhook TEXT')
        except sqlite3.OperationalError:
            pass  # Column already exists
        
        # Initialize default settings
        init_default_settings(conn)
        
        conn.commit()

def ensure_db():
    """Ensure database exists and is initialized"""
    try:
        with get_db_connection() as conn:
            # Test if we can connect and create tables if needed
            init_db()
    except Exception as e:
        print(f"Database initialization error: {e}")

class SSLChecker:
    def __init__(self):
        self.results = {}
    
    def get_ssl_info_by_ip(self, ip, port=443, hostname=None):
        """Get SSL certificate information by IP and hostname"""
        try:
            print(f"🔍 Attempting SSL connection to {ip}:{port} (hostname: {hostname})")
            
            # Create socket connection to IP
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(10)  # Extended timeout - 10 seconds for better reliability
            
            # Connect to IP:port
            sock.connect((ip, port))
            print(f"✅ TCP connection established to {ip}:{port}")
            
            # Create SSL context and wrap socket
            context = ssl.create_default_context()
            # Don't verify hostname for IP connections
            context.check_hostname = False
            context.verify_mode = ssl.CERT_NONE
            
            with context.wrap_socket(sock, server_hostname=hostname) as ssock:
                cert = ssock.getpeercert()
                print(f"✅ SSL handshake successful for {hostname or ip}")
                
                # Try binary format if text format is empty
                if not cert:
                    print("🔄 Trying binary certificate format...")
                    cert_bin = ssock.getpeercert(binary_form=True)
                    if cert_bin:
                        cert_obj = cryptography.x509.load_der_x509_certificate(cert_bin)
                        print(f"✅ Binary certificate parsed successfully")
                        
                        # Extract information from cryptography object
                        expiry_date = cert_obj.not_valid_after_utc.replace(tzinfo=None)
                        days_left = (expiry_date - datetime.datetime.now()).days
                        
                        # Extract issuer and subject common names
                        issuer_cn = "Unknown"
                        for attribute in cert_obj.issuer:
                            if attribute.oid._name == 'commonName':
                                issuer_cn = attribute.value
                                break
                        
                        subject_cn = "Unknown"
                        for attribute in cert_obj.subject:
                            if attribute.oid._name == 'commonName':
                                subject_cn = attribute.value
                                break
                        
                        print(f"✅ Certificate parsed: expires {expiry_date.strftime('%Y-%m-%d')}, {days_left} days left")
                        
                        return {
                            'status': 'success',
                            'issuer': issuer_cn,
                            'subject': subject_cn,
                            'expiry_date': expiry_date.strftime('%Y-%m-%d'),
                            'days_left': days_left,
                            'serial_number': str(cert_obj.serial_number),
                            'version': str(cert_obj.version)
                        }
                    else:
                        print("❌ Binary certificate is also None")
                        return {
                            'status': 'error',
                            'message': 'No certificate found in any format'
                        }
                
                if not cert:
                    return {
                        'status': 'error',
                        'message': 'No certificate found'
                    }
                
                # Parse expiry date
                expiry_str = cert['notAfter']
                expiry_date = datetime.datetime.strptime(expiry_str, '%b %d %H:%M:%S %Y %Z')
                
                # Calculate days until expiry
                days_left = (expiry_date - datetime.datetime.now()).days
                
                print(f"✅ Certificate parsed: expires {expiry_date.strftime('%Y-%m-%d')}, {days_left} days left")
                
                return {
                    'status': 'success',
                    'issuer': cert.get('issuer', [{}])[-1].get('commonName', 'Unknown'),
                    'subject': cert.get('subject', [{}])[-1].get('commonName', 'Unknown'),
                    'expiry_date': expiry_date.strftime('%Y-%m-%d'),
                    'days_left': days_left,
                    'serial_number': cert.get('serialNumber', 'Unknown'),
                    'version': cert.get('version', 'Unknown')
                }
        
        except socket.timeout:
            print(f"❌ Connection timeout to {ip}:{port}")
            return {
                'status': 'error',
                'message': 'Connection timeout (10s) - Server may be down or port blocked'
            }
        except ConnectionRefusedError:
            print(f"❌ Connection refused to {ip}:{port}")
            return {
                'status': 'error',
                'message': 'Connection refused - SSL service not available'
            }
        except ssl.SSLError as e:
            print(f"❌ SSL error for {ip}:{port}: {e}")
            return {
                'status': 'error',
                'message': f'SSL error: {str(e)}'
            }
        except Exception as e:
            print(f"❌ Unexpected error for {ip}:{port}: {e}")
            return {
                'status': 'error',
                'message': f'Unexpected error: {str(e)}'
            }
        except socket.error as e:
            return {
                'status': 'error',
                'message': f'Socket error: {str(e)}'
            }
        except ssl.SSLError as e:
            return {
                'status': 'error',
                'message': f'SSL error: {str(e)}'
            }
        except Exception as e:
            return {
                'status': 'error',
                'message': f'Unexpected error: {str(e)}'
            }

# Initialize SSL checker
ssl_checker = SSLChecker()

@app.route('/login', methods=['GET', 'POST'])
def login():
    # Initialize database on first access
    ensure_db()
    
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        if username == USERNAME and password == PASSWORD:
            session['logged_in'] = True
            return redirect(url_for('index'))
        else:
            flash('Invalid username or password')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.pop('logged_in', None)
    return redirect(url_for('login'))

@app.route('/')
@login_required
def index():
    ensure_db()
    return render_template('index.html')

@app.route('/check-ssl', methods=['POST'])
@login_required
def check_ssl():
    data = request.get_json()
    
    if not data:
        return jsonify({'error': 'No data provided'}), 400
    
    target = data.get('target')
    if not target:
        return jsonify({'error': 'Target is required'}), 400
    
    # Parse hostname:port format
    port = 443  # Default port
    if ':' in target:
        hostname_part, port_part = target.rsplit(':', 1)
        try:
            port = int(port_part)
            target = hostname_part
        except ValueError:
            return jsonify({'error': 'Invalid port number'}), 400
    
    # If it's an IP address, use it directly
    try:
        socket.inet_aton(target)
        ip = target
        hostname = None
    except socket.error:
        # It's a hostname, resolve it
        try:
            ip = socket.gethostbyname(target)
            hostname = target
        except socket.gaierror:
            return jsonify({'error': 'Could not resolve hostname'}), 400
    
    ssl_info = ssl_checker.get_ssl_info_by_ip(ip, port, hostname)
    ssl_info['resolved_ip'] = ip
    ssl_info['port'] = port
    
    return jsonify(ssl_info)

@app.route('/check-bulk-ssl', methods=['POST'])
@login_required
def check_bulk_ssl():
    """Bulk SSL check endpoint for main page"""
    data = request.get_json()
    
    if not data:
        return jsonify({'error': 'No data provided'}), 400
    
    domains_text = data.get('domains')
    if not domains_text:
        return jsonify({'error': 'Domains are required'}), 400
    
    # Parse domains from text
    domain_lines = [line.strip() for line in domains_text.split('\n') if line.strip()]
    
    if len(domain_lines) == 0:
        return jsonify({'error': 'No valid domains provided'}), 400
    
    if len(domain_lines) > 50:
        return jsonify({'error': 'Maximum 50 domains allowed'}), 400
    
    # Generate session ID for tracking
    session_id = str(uuid.uuid4())
    
    def run_bulk_check():
        """Background thread function for bulk SSL checking"""
        results = []
        total_domains = len(domain_lines)
        
        for i, domain_line in enumerate(domain_lines):
            # Parse hostname:port format
            port = 443  # Default port
            target = domain_line
            
            if ':' in domain_line:
                hostname_part, port_part = domain_line.rsplit(':', 1)
                try:
                    port = int(port_part)
                    target = hostname_part
                except ValueError:
                    port = 443  # Keep default if invalid port
            
            # Resolve hostname to IP
            try:
                # Check if it's already an IP address
                try:
                    socket.inet_aton(target)
                    ip = target
                    hostname = None
                except socket.error:
                    # It's a hostname, resolve it
                    ip = socket.gethostbyname(target)
                    hostname = target
                
                # Check SSL
                ssl_info = ssl_checker.get_ssl_info_by_ip(ip, port, hostname)
                
                # Organize data for export
                result_data = {
                    'domain': domain_line,
                    'resolved_ip': ip,
                    'port': port,
                    'status': ssl_info.get('status', 'error'),
                    'message': ssl_info.get('message', ''),
                    'ssl_info': {
                        'valid': ssl_info.get('status') == 'success',
                        'issuer': ssl_info.get('issuer', ''),
                        'subject': ssl_info.get('subject', ''),
                        'expire_date': ssl_info.get('expiry_date', ''),
                        'days_left': ssl_info.get('days_left', ''),
                        'serial_number': ssl_info.get('serial_number', '')
                    }
                }
                
                results.append(result_data)
                
            except socket.gaierror:
                # DNS resolution failed
                results.append({
                    'domain': domain_line,
                    'status': 'error',
                    'message': 'Could not resolve hostname',
                    'resolved_ip': None,
                    'port': port
                })
            except Exception as e:
                # Other errors
                results.append({
                    'domain': domain_line,
                    'status': 'error',
                    'message': str(e),
                    'resolved_ip': None,
                    'port': port
                })
            
            # Update progress
            progress = int((i + 1) / total_domains * 100)
            is_completed = i + 1 == total_domains
            
            ssl_checker.results[session_id] = {
                'progress': progress,
                'results': results,
                'completed': is_completed,
                'total': total_domains,
                'processed': i + 1
            }
    
    # Start background thread
    thread = threading.Thread(target=run_bulk_check)
    thread.daemon = True
    thread.start()
    
    return jsonify({'session_id': session_id})

@app.route('/settings')
@login_required
def settings_page():
    """Settings management page"""
    ensure_db()
    settings = get_all_settings()
    return render_template('settings.html', settings=settings)

@app.route('/api/settings', methods=['GET'])
@login_required
def get_settings_api():
    """Get all settings API"""
    settings = get_all_settings()
    return jsonify(settings)

@app.route('/api/settings', methods=['POST'])
@login_required
def update_settings_api():
    """Update settings API"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        updated_count = 0
        with get_db_connection() as conn:
            for key, value in data.items():
                # Convert boolean values to string
                if isinstance(value, bool):
                    value = 'true' if value else 'false'
                
                conn.execute('''
                    UPDATE settings SET value = ?, updated_at = ?
                    WHERE key = ?
                ''', (str(value), datetime.datetime.now().isoformat(), key))
                updated_count += 1
            
            conn.commit()
        
        return jsonify({
            'success': True, 
            'message': f'{updated_count} settings updated successfully'
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/settings/test-email', methods=['POST'])
@login_required
def test_email_settings():
    """Test email settings"""
    try:
        data = request.get_json()
        test_email = data.get('test_email')
        
        if not test_email:
            return jsonify({'error': 'Test email address required'}), 400
        
        # Get current SMTP settings from database
        smtp_server = get_setting('smtp_server')
        smtp_port_str = get_setting('smtp_port', '587')
        smtp_username = get_setting('smtp_username')
        smtp_password = get_setting('smtp_password')
        from_email = get_setting('smtp_from_email')
        use_tls = get_setting('smtp_use_tls', 'true').lower() == 'true'
        
        # Validate settings
        missing_settings = []
        if not smtp_server or smtp_server.strip() == '':
            missing_settings.append('SMTP Server')
        if not smtp_username or smtp_username.strip() == '':
            missing_settings.append('SMTP Username')
        if not smtp_password or smtp_password.strip() == '':
            missing_settings.append('SMTP Password')
        if not from_email or from_email.strip() == '':
            missing_settings.append('From Email')
            
        if missing_settings:
            return jsonify({
                'error': f'SMTP settings are incomplete. Missing: {", ".join(missing_settings)}'
            }), 400
        
        try:
            smtp_port = int(smtp_port_str)
        except (ValueError, TypeError):
            return jsonify({'error': 'Invalid SMTP port number'}), 400
        
        # Send test email
        msg = email.mime.multipart.MIMEMultipart()
        msg['From'] = from_email
        msg['To'] = test_email
        msg['Subject'] = 'SSL Checker - Test Email'
        
        body = """
        This is a test email from SSL Checker.
        
        If you received this email, your SMTP configuration is working correctly.
        
        Test sent at: """ + datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        msg.attach(email.mime.text.MIMEText(body, 'plain'))
        
        server = smtplib.SMTP(smtp_server, smtp_port)
        if use_tls:
            server.starttls()
        server.login(smtp_username, smtp_password)
        server.send_message(msg)
        server.quit()
        
        return jsonify({
            'success': True,
            'message': f'Test email sent successfully to {test_email}'
        })
        
    except Exception as e:
        return jsonify({'error': f'Email test failed: {str(e)}'}), 500

@app.route('/api/settings/test-slack', methods=['POST'])
@login_required
def test_slack_settings():
    """Test Slack settings"""
    try:
        data = request.get_json()
        webhook_url = data.get('webhook_url') or get_setting('default_slack_webhook')
        
        if not webhook_url:
            return jsonify({'error': 'Slack webhook URL required'}), 400
        
        if 'YOUR/SLACK/WEBHOOK' in webhook_url:
            return jsonify({'error': 'Please configure a real Slack webhook URL'}), 400
        
        # Send test message
        message = {
            "text": f"🧪 SSL Checker Test Message - {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        }
        
        response = requests.post(webhook_url, json=message, timeout=10)
        
        if response.status_code == 200:
            return jsonify({
                'success': True,
                'message': 'Test message sent successfully to Slack'
            })
        else:
            return jsonify({
                'error': f'Slack test failed: HTTP {response.status_code} - {response.text}'
            }), 400
        
    except Exception as e:
        return jsonify({'error': f'Slack test failed: {str(e)}'}), 500

@app.route('/expire-tracking')
@login_required
def expire_tracking():
    with get_db_connection() as conn:
        domains_raw = conn.execute('''
            SELECT d.*, 
                   COUNT(s.id) as subdomain_count,
                   COUNT(CASE WHEN s.certificate_renewal_status = 'Valid' THEN 1 END) as valid_count,
                   COUNT(CASE WHEN s.certificate_renewal_status = 'Warning' THEN 1 END) as warning_count,
                   COUNT(CASE WHEN s.certificate_renewal_status = 'Expired' THEN 1 END) as expired_count,
                   COUNT(CASE WHEN s.certificate_renewal_status = 'Error' OR s.certificate_renewal_status IS NULL THEN 1 END) as error_count
            FROM domain d
            LEFT JOIN subdomain s ON d.id = s.domain_id
            GROUP BY d.id
            ORDER BY d.created_at DESC
        ''').fetchall()
    
    # Convert created_at strings to datetime objects
    domains = []
    for domain in domains_raw:
        domain_dict = dict(domain)
        if domain_dict['created_at']:
            try:
                # Try ISO format first
                domain_dict['created_at'] = datetime.datetime.fromisoformat(domain_dict['created_at'].replace('Z', '+00:00'))
            except:
                try:
                    # Try standard SQL format
                    domain_dict['created_at'] = datetime.datetime.strptime(domain_dict['created_at'], '%Y-%m-%d %H:%M:%S')
                except:
                    try:
                        # Try date only format
                        domain_dict['created_at'] = datetime.datetime.strptime(domain_dict['created_at'], '%Y-%m-%d')
                    except:
                        print(f"Warning: Could not parse date {domain_dict['created_at']} for domain {domain_dict['name']}")
                        domain_dict['created_at'] = None
        else:
            domain_dict['created_at'] = None
        
        domains.append(domain_dict)
    
    return render_template('expire_tracking.html', domains=domains)

@app.route('/expire-tracking/domain/<int:domain_id>')
def domain_detail(domain_id):
    with get_db_connection() as conn:
        domain_raw = conn.execute('SELECT * FROM domain WHERE id = ?', (domain_id,)).fetchone()
        if not domain_raw:
            return "Domain not found", 404
        
        subdomains_raw = conn.execute('''
            SELECT * FROM subdomain 
            WHERE domain_id = ? 
            ORDER BY dns
        ''', (domain_id,)).fetchall()
    
    # Convert domain created_at to datetime object
    domain = dict(domain_raw)
    if domain['created_at']:
        try:
            domain['created_at'] = datetime.datetime.fromisoformat(domain['created_at'].replace('Z', '+00:00'))
        except:
            try:
                domain['created_at'] = datetime.datetime.strptime(domain['created_at'], '%Y-%m-%d %H:%M:%S')
            except:
                domain['created_at'] = datetime.datetime.now()
    
    # Convert subdomain dates to datetime objects
    subdomains = []
    for subdomain in subdomains_raw:
        subdomain_dict = dict(subdomain)
        
        # Convert last_checked to datetime object
        if subdomain_dict['last_checked']:
            try:
                subdomain_dict['last_checked'] = datetime.datetime.fromisoformat(subdomain_dict['last_checked'].replace('Z', '+00:00'))
            except:
                try:
                    subdomain_dict['last_checked'] = datetime.datetime.strptime(subdomain_dict['last_checked'], '%Y-%m-%d %H:%M:%S')
                except:
                    subdomain_dict['last_checked'] = None
        
        # Convert expire_date to date object
        if subdomain_dict['expire_date']:
            try:
                subdomain_dict['expire_date'] = datetime.datetime.strptime(subdomain_dict['expire_date'], '%Y-%m-%d').date()
            except:
                subdomain_dict['expire_date'] = None
        
        subdomains.append(subdomain_dict)
    
    return render_template('domain_detail.html', domain=domain, subdomains=subdomains)

@app.route('/expire-tracking/add-domain', methods=['POST'])
def add_domain():
    data = request.get_json()
    domain_name = data.get('name')
    
    if not domain_name:
        return jsonify({'error': 'Domain name is required'}), 400
    
    try:
        with get_db_connection() as conn:
            cursor = conn.execute(
                'INSERT INTO domain (name) VALUES (?)',
                (domain_name,)
            )
            conn.commit()
            domain_id = cursor.lastrowid
        
        return jsonify({'success': True, 'domain_id': domain_id})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/expire-tracking/domain/<int:domain_id>/add-subdomain', methods=['POST'])
def add_subdomain(domain_id):
    data = request.get_json()
    dns = data.get('dns')
    ip = data.get('ip')
    port = data.get('port', 443)
    
    if not dns or not ip:
        return jsonify({'error': 'DNS and IP are required'}), 400
    
    try:
        with get_db_connection() as conn:
            cursor = conn.execute('''
                INSERT INTO subdomain (domain_id, dns, ip, port)
                VALUES (?, ?, ?, ?)
            ''', (domain_id, dns, ip, port))
            conn.commit()
            subdomain_id = cursor.lastrowid
        
        return jsonify({'success': True, 'subdomain_id': subdomain_id})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/expire-tracking/subdomain/<int:subdomain_id>/check-ssl', methods=['POST'])
def check_subdomain_ssl(subdomain_id):
    data = request.get_json() or {}
    use_nat = data.get('use_nat', False)
    
    with get_db_connection() as conn:
        subdomain = conn.execute('SELECT * FROM subdomain WHERE id = ?', (subdomain_id,)).fetchone()
        if not subdomain:
            return jsonify({'error': 'Subdomain not found'}), 404
    
    # SSL Check Logic:
    # 1. NAT SSL Check -> Use NAT IP
    # 2. Proxied Yes -> Use domain name (DNS resolution)
    # 3. Proxied No -> Use stored IP address
    
    hostname = subdomain['dns']
    port = subdomain['port']
    
    if use_nat and subdomain['nat']:
        # NAT SSL Check - use NAT IP
        target_ip = str(subdomain['nat']).strip()
        print(f"🔧 NAT SSL Check: {hostname} via NAT IP {target_ip}:{port}")
    elif subdomain['proxied'] == 'Yes':
        # Proxied Yes - use domain name (resolve DNS in real-time)
        try:
            target_ip = socket.gethostbyname(subdomain['dns'])
            print(f"🌐 Proxied SSL Check: {hostname} resolved to {target_ip}:{port}")
        except socket.gaierror:
            return jsonify({'error': f'Cannot resolve DNS for {subdomain["dns"]}'}), 400
    else:
        # Proxied No or empty - use stored IP address
        target_ip = str(subdomain['ip']).strip()
        if not target_ip or target_ip == 'None':
            return jsonify({'error': f'No IP address stored for {subdomain["dns"]}'}), 400
        print(f"🎯 Direct IP SSL Check: {hostname} via stored IP {target_ip}:{port}")
    
    try:
        ssl_info = ssl_checker.get_ssl_info_by_ip(target_ip, port, hostname)
        
        if ssl_info['status'] == 'success':
            # Update database
            with get_db_connection() as conn:
                conn.execute('''
                    UPDATE subdomain SET 
                    certificate = ?, expire_date = ?, days_left = ?, 
                    certificate_renewal_status = ?, last_checked = ?
                    WHERE id = ?
                ''', (
                    ssl_info['issuer'],
                    ssl_info['expiry_date'],
                    ssl_info['days_left'],
                    'Valid' if ssl_info['days_left'] > 30 else 'Warning' if ssl_info['days_left'] > 0 else 'Expired',
                    datetime.datetime.now().isoformat(),
                    subdomain_id
                ))
                conn.commit()
        
        return jsonify(ssl_info)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/expire-tracking/domain/<int:domain_id>/check-all-ssl', methods=['POST'])
def check_all_domain_ssl(domain_id):
    print(f"🚀 check_all_domain_ssl called for domain_id: {domain_id}")
    
    with get_db_connection() as conn:
        domain = conn.execute('SELECT * FROM domain WHERE id = ?', (domain_id,)).fetchone()
        if not domain:
            print(f"❌ Domain {domain_id} not found")
            return jsonify({'error': 'Domain not found'}), 404
        
        subdomains = conn.execute('SELECT * FROM subdomain WHERE domain_id = ?', (domain_id,)).fetchall()
        print(f"📊 Found {len(subdomains)} subdomains for domain {domain['name']}")
    
    session_id = str(uuid.uuid4())
    session['current_domain_check'] = session_id
    print(f"🎯 Generated session_id: {session_id}")
    
    def run_domain_check():
        results = []
        total_subdomains = len(subdomains)
        print(f"Starting PARALLEL domain check for {total_subdomains} subdomains, session: {session_id}")
        
        def check_single_subdomain(index_subdomain):
            i, subdomain = index_subdomain
            print(f"Checking subdomain {i+1}/{total_subdomains}: {subdomain['dns']}")
            
            # Apply same logic as individual SSL check:
            # 1. Proxied Yes -> Use domain name (DNS resolution)
            # 2. Proxied No -> Use stored IP address
            
            hostname = subdomain['dns']
            port = subdomain['port']
            
            if subdomain['proxied'] == 'Yes':
                # Proxied Yes - use domain name (resolve DNS in real-time)
                try:
                    target_ip = socket.gethostbyname(subdomain['dns'])
                    print(f"🌐 Proxied SSL Check: {hostname} resolved to {target_ip}:{port}")
                except socket.gaierror:
                    ssl_info = {'status': 'error', 'message': f'Cannot resolve DNS for {subdomain["dns"]}'}
                    print(f"❌ DNS resolution failed for {subdomain['dns']}")
                    target_ip = None
            else:
                # Proxied No or empty - use stored IP address
                target_ip = str(subdomain['ip']).strip() if subdomain['ip'] else None
                if not target_ip or target_ip == 'None':
                    ssl_info = {'status': 'error', 'message': f'No IP address stored for {subdomain["dns"]}'}
                    print(f"❌ No IP address for {subdomain['dns']}")
                    target_ip = None
                else:
                    print(f"🎯 Direct IP SSL Check: {hostname} via stored IP {target_ip}:{port}")
            
            if target_ip:
                ssl_info = ssl_checker.get_ssl_info_by_ip(target_ip, port, hostname)
                print(f"SSL result for {subdomain['dns']}: {ssl_info}")
                
                if ssl_info['status'] == 'success':
                    with get_db_connection() as conn:
                        conn.execute('''
                            UPDATE subdomain SET 
                            certificate = ?, expire_date = ?, days_left = ?, 
                            certificate_renewal_status = ?, last_checked = ?
                            WHERE id = ?
                        ''', (
                            ssl_info['issuer'],
                            ssl_info['expiry_date'],
                            ssl_info['days_left'],
                            'Valid' if ssl_info['days_left'] > 30 else 'Warning' if ssl_info['days_left'] > 0 else 'Expired',
                            datetime.datetime.now().isoformat(),
                            subdomain['id']
                        ))
                        conn.commit()
                        print(f"✅ Updated subdomain {subdomain['dns']} in database - Days left: {ssl_info['days_left']}")
                else:
                    # SSL check failed, still update last_checked
                    with get_db_connection() as conn:
                        conn.execute('''
                            UPDATE subdomain SET 
                            certificate_renewal_status = ?, last_checked = ?
                            WHERE id = ?
                        ''', (
                            'Error',
                            datetime.datetime.now().isoformat(),
                            subdomain['id']
                        ))
                        conn.commit()
                        print(f"❌ SSL check failed for {subdomain['dns']}: {ssl_info.get('message', 'Unknown error')}")
            else:
                # Error case already handled above
                with get_db_connection() as conn:
                    conn.execute('''
                        UPDATE subdomain SET 
                        certificate_renewal_status = ?, last_checked = ?
                        WHERE id = ?
                    ''', (
                        'Error',
                        datetime.datetime.now().isoformat(),
                        subdomain['id']
                    ))
                    conn.commit()
            
            return {
                'subdomain_id': subdomain['id'],
                'dns': subdomain['dns'],
                'result': ssl_info,
                'index': i
            }
        
        # Use ThreadPoolExecutor for parallel processing (max 5 concurrent)
        completed_count = 0
        with ThreadPoolExecutor(max_workers=5) as executor:
            # Submit all tasks
            future_to_subdomain = {
                executor.submit(check_single_subdomain, (i, subdomain)): i 
                for i, subdomain in enumerate(subdomains)
            }
            
            # Process completed tasks as they finish
            for future in as_completed(future_to_subdomain):
                try:
                    result = future.result()
                    results.append(result)
                    completed_count += 1
                    
                    progress = int(completed_count / total_subdomains * 100)
                    is_completed = completed_count == total_subdomains
                    
                    ssl_checker.results[session_id] = {
                        'progress': progress,
                        'results': sorted(results, key=lambda x: x['index']),  # Keep order
                        'completed': is_completed
                    }
                    
                    print(f"Progress: {progress}% ({completed_count}/{total_subdomains}), Completed: {is_completed}")
                    
                except Exception as exc:
                    print(f"❌ Subdomain check generated an exception: {exc}")
                    completed_count += 1
        
        print(f"PARALLEL domain check completed for session: {session_id}")
    
    thread = threading.Thread(target=run_domain_check)
    thread.daemon = True
    thread.start()
    
    return jsonify({'session_id': session_id})

@app.route('/expire-tracking/domain/<int:domain_id>/check-progress/<session_id>')
def check_domain_ssl_progress(domain_id, session_id):
    if session_id in ssl_checker.results:
        return jsonify(ssl_checker.results[session_id])
    else:
        return jsonify({'progress': 0, 'results': [], 'completed': False})

@app.route('/progress/<session_id>')
def check_progress(session_id):
    """General progress endpoint for SSL checking sessions"""
    print(f"Progress check requested for session: {session_id}")
    if session_id in ssl_checker.results:
        result = ssl_checker.results[session_id]
        print(f"Progress found: {result['progress']}%, completed: {result['completed']}")
        return jsonify(result)
    else:
        print(f"Session {session_id} not found in results")
        return jsonify({'progress': 0, 'results': [], 'completed': False})

@app.route('/expire-tracking/domain/<int:domain_id>/export')
def export_excel(domain_id):
    try:
        from openpyxl import Workbook
        from flask import send_file
        import io
        
        with get_db_connection() as conn:
            domain = conn.execute('SELECT * FROM domain WHERE id = ?', (domain_id,)).fetchone()
            if not domain:
                return "Domain not found", 404
            
            subdomains = conn.execute('''
                SELECT * FROM subdomain 
                WHERE domain_id = ? 
                ORDER BY dns
            ''', (domain_id,)).fetchall()
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"{domain['name']}_SSL_Report"
        
        # Headers
        headers = [
            'DNS', 'Source DNS', 'Proxied', 'Record Type', 'IP', 'NAT', 'Port', 'Owner',
            'Certificate', 'Expire Date', 'Days Left', 'Status', 'Last Checked'
        ]
        ws.append(headers)
        
        # Data
        for subdomain in subdomains:
            row = [
                subdomain['dns'] or '',
                subdomain['source_dns'] or '',
                subdomain['proxied'] or '',
                subdomain['record_type'] or 'A',
                subdomain['ip'] or '',
                subdomain['nat'] or '',
                subdomain['port'] or 443,
                subdomain['owner'] or '',
                subdomain['certificate'] or '',
                subdomain['expire_date'] or '',
                subdomain['days_left'] or '',
                subdomain['certificate_renewal_status'] or '',
                subdomain['last_checked'] or ''
            ]
            ws.append(row)
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'{domain["name"]}_ssl_report.xlsx'
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/expire-tracking/domain/<int:domain_id>/import-excel', methods=['POST'])
def import_excel(domain_id):
    with get_db_connection() as conn:
        domain = conn.execute('SELECT * FROM domain WHERE id = ?', (domain_id,)).fetchone()
        if not domain:
            return jsonify({'error': 'Domain not found'}), 404
    
    if 'file' not in request.files:
        return jsonify({'error': 'Dosya seçilmedi'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Dosya seçilmedi'}), 400
    
    try:
        # Excel dosyasını oku
        wb = load_workbook(file)
        ws = wb.active
        
        imported_count = 0
        with get_db_connection() as conn:
            for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header
                if row[0]:  # DNS column not empty
                    conn.execute('''
                        INSERT INTO subdomain 
                        (domain_id, dns, source_dns, proxied, record_type, ip, nat, port, owner)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (
                        domain_id,
                        str(row[0] or ''),
                        str(row[1] or ''),
                        str(row[2] or ''),
                        str(row[3] or 'A'),
                        str(row[4] or ''),
                        str(row[5] or ''),
                        int(row[6] or 443),
                        str(row[7] or '')
                    ))
                    imported_count += 1
            
            conn.commit()
        
        return jsonify({'success': True, 'imported_count': imported_count})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/expire-tracking/subdomain/<int:subdomain_id>/update-settings', methods=['POST'])
@login_required
def update_subdomain_settings(subdomain_id):
    """Update subdomain-specific notification settings"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        custom_email = data.get('custom_email', '').strip()
        custom_slack_webhook = data.get('custom_slack_webhook', '').strip()
        
        # Convert empty strings to None for database
        custom_email = custom_email if custom_email else None
        custom_slack_webhook = custom_slack_webhook if custom_slack_webhook else None
        
        with get_db_connection() as conn:
            conn.execute('''
                UPDATE subdomain SET 
                custom_email = ?, custom_slack_webhook = ?
                WHERE id = ?
            ''', (custom_email, custom_slack_webhook, subdomain_id))
            conn.commit()
        
        return jsonify({'success': True, 'message': 'Subdomain settings updated successfully'})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/expire-tracking/subdomain/<int:subdomain_id>/update', methods=['PUT'])
def update_subdomain(subdomain_id):
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No JSON data provided'}), 400
        
        with get_db_connection() as conn:
            conn.execute('''
                UPDATE subdomain SET 
                dns = ?, source_dns = ?, proxied = ?, record_type = ?,
                ip = ?, nat = ?, port = ?, owner = ?
                WHERE id = ?
            ''', (
                data.get('dns'),
                data.get('source_dns'),
                data.get('proxied'),
                data.get('record_type'),
                data.get('ip'),
                data.get('nat'),
                data.get('port'),
                data.get('owner'),
                subdomain_id
            ))
            conn.commit()
        
        return jsonify({'success': True})
    except Exception as e:
        print(f"❌ Update subdomain error: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/expire-tracking/subdomain/<int:subdomain_id>/delete', methods=['DELETE'])
def delete_subdomain(subdomain_id):
    try:
        with get_db_connection() as conn:
            conn.execute('DELETE FROM subdomain WHERE id = ?', (subdomain_id,))
            conn.commit()
        
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/expire-tracking/domain/<int:domain_id>/delete', methods=['DELETE'])
def delete_domain(domain_id):
    try:
        with get_db_connection() as conn:
            # Delete subdomains first
            conn.execute('DELETE FROM subdomain WHERE domain_id = ?', (domain_id,))
            # Delete domain
            conn.execute('DELETE FROM domain WHERE id = ?', (domain_id,))
            conn.commit()
        
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/expire-tracking/subdomain/<int:subdomain_id>/toggle-notification', methods=['POST'])
def toggle_notification(subdomain_id):
    data = request.get_json()
    notification_type = data.get('type')  # 'email' or 'slack'
    enabled = data.get('enabled', False)
    
    try:
        column_name = f"{notification_type}_notification"
        with get_db_connection() as conn:
            conn.execute(f'''
                UPDATE subdomain SET {column_name} = ?
                WHERE id = ?
            ''', (1 if enabled else 0, subdomain_id))
            conn.commit()
        
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/expire-tracking/subdomain/<int:subdomain_id>/send-notification', methods=['POST'])
@login_required
def send_manual_notification(subdomain_id):
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
            
        notification_type = data.get('type')  # 'email' or 'slack'
        if not notification_type:
            return jsonify({'error': 'Notification type not specified'}), 400
        
        ensure_db()
        with get_db_connection() as conn:
            subdomain = conn.execute('SELECT * FROM subdomain WHERE id = ?', (subdomain_id,)).fetchone()
            if not subdomain:
                return jsonify({'error': 'Subdomain not found'}), 404
        
        subdomain_dict = dict(subdomain)
        
        if notification_type == 'email':
            success = send_email_notification(subdomain_dict)
            message = 'Email notification sent successfully!' if success else 'Email notification failed to send'
        elif notification_type == 'slack':
            success = send_slack_notification(subdomain_dict)
            message = 'Slack notification sent successfully!' if success else 'Slack notification failed to send. Check webhook URL in settings'
        else:
            return jsonify({'error': 'Invalid notification type'}), 400
        
        return jsonify({
            'success': success,
            'message': message,
            'type': notification_type
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Notification functions
def send_email_notification(subdomain, config=None):
    """Send email notification for SSL expiry"""
    try:
        # Get email settings from database
        smtp_server = get_setting('smtp_server')
        smtp_port = int(get_setting('smtp_port', '587'))
        smtp_username = get_setting('smtp_username')
        smtp_password = get_setting('smtp_password')
        from_email = get_setting('smtp_from_email')
        use_tls = get_setting('smtp_use_tls', 'true').lower() == 'true'
        email_enabled = get_setting('email_enabled', 'true').lower() == 'true'
        
        if not email_enabled:
            return False
            
        if not all([smtp_server, smtp_username, smtp_password, from_email]):
            print("Email settings incomplete")
            return False
        
        # Use custom email for subdomain if set, otherwise use default
        to_email = subdomain.get('custom_email') or get_setting('default_to_email')
        if not to_email:
            print("No recipient email configured")
            return False
            
        msg = email.mime.multipart.MIMEMultipart()
        msg['From'] = from_email
        msg['To'] = to_email
        msg['Subject'] = f"SSL Certificate Expiry Alert - {subdomain['dns']}"
        
        body = f"""
        SSL Certificate Expiry Alert
        
        Domain: {subdomain['dns']}
        Days Left: {subdomain.get('days_left', 'Unknown')}
        Expiry Date: {subdomain.get('expire_date', 'Unknown')}
        Certificate Issuer: {subdomain.get('certificate', 'Unknown')}
        
        Please renew the SSL certificate before it expires.
        
        Generated at: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
        """
        
        msg.attach(email.mime.text.MIMEText(body, 'plain'))
        
        server = smtplib.SMTP(smtp_server, smtp_port)
        if use_tls:
            server.starttls()
        server.login(smtp_username, smtp_password)
        server.send_message(msg)
        server.quit()
        
        return True
    except Exception as e:
        print(f"Email send error: {e}")
        return False

def send_slack_notification(subdomain, config=None):
    """Send Slack notification for SSL expiry"""
    try:
        slack_enabled = get_setting('slack_enabled', 'true').lower() == 'true'
        if not slack_enabled:
            print("Slack notifications are disabled")
            return False
        
        # Use custom webhook for subdomain if set, otherwise use default
        webhook_url = subdomain.get('custom_slack_webhook') or get_setting('default_slack_webhook')
        if not webhook_url:
            print("No Slack webhook URL configured")
            return False
            
        # Check if webhook URL is valid (not placeholder)
        if 'YOUR/SLACK/WEBHOOK' in webhook_url:
            print("Slack webhook URL is still a placeholder - please configure a real webhook URL")
            return False
            
        days_left = subdomain.get('days_left', 0)
        emoji = "🔴" if days_left <= 7 else "🟡" if days_left <= 30 else "🟢"
        
        # Get Slack settings
        channel = get_setting('slack_channel', '#ssl-alerts')
        username = get_setting('slack_username', 'SSL Checker Bot')
        icon_emoji = get_setting('slack_icon_emoji', ':warning:')
        
        # Enhanced Slack message
        message = {
            "text": f"{emoji} SSL Certificate Alert",
            "username": username,
            "icon_emoji": icon_emoji,
            "channel": channel,
            "attachments": [
                {
                    "color": "danger" if days_left <= 7 else "warning" if days_left <= 30 else "good",
                    "fields": [
                        {
                            "title": "Domain",
                            "value": subdomain['dns'],
                            "short": True
                        },
                        {
                            "title": "Days Left",
                            "value": str(days_left),
                            "short": True
                        },
                        {
                            "title": "Expiry Date",
                            "value": subdomain.get('expire_date', 'Unknown'),
                            "short": True
                        },
                        {
                            "title": "Certificate Issuer",
                            "value": subdomain.get('certificate', 'Unknown'),
                            "short": True
                        }
                    ]
                }
            ]
        }
        
        print(f"Sending Slack notification to: {webhook_url[:50]}...")
        response = requests.post(webhook_url, json=message, timeout=10)
        
        print(f"Slack response status: {response.status_code}")
        
        if response.status_code == 200:
            print("✅ Slack notification sent successfully")
            return True
        else:
            print(f"❌ Slack notification failed: HTTP {response.status_code} - {response.text}")
            return False
            
    except requests.exceptions.RequestException as e:
        print(f"❌ Slack request error: {e}")
        return False
    except Exception as e:
        print(f"❌ Slack send error: {e}")
        return False

def check_and_send_notifications():
    """Scheduled job to check SSL certificates and send notifications"""
    try:
        # Get notification settings from database
        notifications_enabled = get_setting('notifications_enabled', 'true').lower() == 'true'
        if not notifications_enabled:
            print("Notifications are disabled")
            return
        
        # Get expiry thresholds
        thresholds_str = get_setting('expiry_thresholds', '7,15,30')
        try:
            thresholds = [int(x.strip()) for x in thresholds_str.split(',') if x.strip()]
        except:
            thresholds = [7, 15, 30]  # fallback
        
        print(f"Checking SSL certificates with thresholds: {thresholds}")
        
        # Get all subdomains that need SSL expiry alerts
        with get_db_connection() as conn:
            for threshold in thresholds:
                # Find subdomains expiring in threshold days
                subdomains = conn.execute('''
                    SELECT * FROM subdomain 
                    WHERE days_left IS NOT NULL 
                    AND days_left <= ? 
                    AND days_left >= 0
                ''', (threshold,)).fetchall()
                
                for subdomain in subdomains:
                    subdomain_dict = dict(subdomain)
                    
                    # Send email if enabled for this subdomain
                    if subdomain['email_notification']:
                        success = send_email_notification(subdomain_dict)
                        if success:
                            print(f"✅ Email sent for {subdomain['dns']} ({subdomain['days_left']} days left)")
                        else:
                            print(f"❌ Email failed for {subdomain['dns']}")
                    
                    # Send Slack if enabled for this subdomain
                    if subdomain['slack_notification']:
                        success = send_slack_notification(subdomain_dict)
                        if success:
                            print(f"✅ Slack sent for {subdomain['dns']} ({subdomain['days_left']} days left)")
                        else:
                            print(f"❌ Slack failed for {subdomain['dns']}")
        
        print(f"Notification check completed at {datetime.datetime.now()}")
        
    except Exception as e:
        print(f"Notification check error: {e}")

def run_scheduler():
    """Run the scheduler in a separate thread"""
    while True:
        schedule.run_pending()
        time.sleep(60)  # Check every minute

def setup_scheduler():
    """Setup scheduled jobs"""
    try:
        notifications_enabled = get_setting('notifications_enabled', 'true').lower() == 'true'
        
        if notifications_enabled:
            schedule_time = get_setting('schedule_time', '09:00')
            schedule.every().day.at(schedule_time).do(check_and_send_notifications)
            
            # Start scheduler in background thread
            scheduler_thread = threading.Thread(target=run_scheduler)
            scheduler_thread.daemon = True
            scheduler_thread.start()
            
            print(f"Notification scheduler started - will run daily at {schedule_time}")
        else:
            print("Notification scheduler disabled")
    except Exception as e:
        print(f"Scheduler setup error: {e}")

@app.route('/export')
@login_required
def export_bulk_ssl():
    """Export bulk SSL checker results to Excel"""
    try:
        from openpyxl import Workbook
        from flask import send_file
        import io
        
        # Get current session data from request parameters
        session_id = request.args.get('session_id')
        
        if not session_id or session_id not in ssl_checker.results:
            return jsonify({'error': 'No data available for export'}), 400
        
        session_data = ssl_checker.results[session_id]
        
        if not session_data.get('completed', False):
            return jsonify({'error': 'SSL check not completed yet'}), 400
        
        # Get results data - using 'results' key not 'domains'
        domains_data = session_data.get('results', [])
        
        if not domains_data:
            return jsonify({'error': 'No domains found in session data'}), 400
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Bulk_SSL_Report"
        
        # Headers
        headers = [
            'Domain', 'Port', 'Status', 'Certificate Valid', 'Issuer', 
            'Expire Date', 'Days Left', 'Subject', 'Serial Number', 'Resolved IP', 'Message'
        ]
        ws.append(headers)
        
        # Data
        for domain_info in domains_data:
            ssl_info = domain_info.get('ssl_info', {})
            row = [
                domain_info.get('domain', ''),
                domain_info.get('port', 443),
                domain_info.get('status', ''),
                'Yes' if ssl_info.get('valid') else 'No',
                ssl_info.get('issuer', ''),
                ssl_info.get('expire_date', ''),
                ssl_info.get('days_left', ''),
                ssl_info.get('subject', ''),
                ssl_info.get('serial_number', ''),
                domain_info.get('resolved_ip', ''),
                domain_info.get('message', '')
            ]
            ws.append(row)
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'bulk_ssl_report_{timestamp}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"Export error: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/health')
def health_check():
    """Health check endpoint for Docker"""
    return jsonify({'status': 'healthy', 'timestamp': datetime.datetime.now().isoformat()})

@app.route('/test-ssl/<host>')
def test_ssl_endpoint(host):
    """Quick SSL test endpoint for debugging"""
    try:
        # Basic validation
        if not host or len(host) > 100:
            return jsonify({'error': 'Invalid host'}), 400
            
        # Try to resolve if it's a domain
        try:
            ip = socket.gethostbyname(host)
        except socket.gaierror:
            # If it fails, assume it's already an IP
            ip = host
            
        result = ssl_checker.get_ssl_info_by_ip(ip, 443, host)
        return jsonify({
            'host': host,
            'ip': ip,
            'result': result
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    # Create instance directory
    os.makedirs('instance', exist_ok=True)
    
    # Initialize database and default settings
    with app.app_context():
        init_db()
        init_default_settings()
    
    # Setup scheduled notifications
    setup_scheduler()
    
    app.run(host='0.0.0.0', port=5000, debug=False)
